"""
export.py — write leads.xlsx from intel.db.

Produces an Excel workbook matching the Bastak lead format:
  * one ALL sheet (every stored item)
  * one sheet per country
  * a SOURCES sheet (the registry) and a LOG sheet (per-run health check)

Formatting follows the house standard: header fill #32A337, white bold header
font, frozen header row, auto column widths.

Phase 1 note: the Score / Company / Project Type / (AI) Summary columns come
from classify.py, which is not built yet — they stay blank until Phase 2.
Rows are sorted so that (once scored) the highest-score leads sit at the top.

Usage:
    python export.py
"""

import re
import sqlite3
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from init_db import DB_PATH, ensure_columns

FRESH_DAYS = 90   # "Fresh" sheet = items published within this many days

SOURCES_PATH = Path(__file__).with_name("sources.yaml")
OUT_PATH = Path(__file__).with_name("leads.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="32A337")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
MAX_WIDTH = 60
MIN_WIDTH = 10

_TAG_RE = re.compile(r"<[^>]+>")


def clean(text: str | None) -> str:
    """Strip HTML tags / collapse whitespace from raw feed summaries."""
    if not text:
        return ""
    text = _TAG_RE.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def sheet_safe(name: str) -> str:
    """Excel sheet-name rules: <=31 chars, none of : \\ / ? * [ ]."""
    name = re.sub(r"[:\\/?*\[\]]", "_", name or "unknown")
    return name[:31] or "unknown"


# (header, function producing the cell value from a sqlite Row)
COLUMNS = [
    ("Published Date", lambda r: r["published_date"] or "unknown"),
    ("Date Found",   lambda r: (r["collected_at"] or "")[:10]),
    ("Score",        lambda r: r["score"]),
    ("Company",      lambda r: r["company"]),
    ("Country",      lambda r: r["country"]),        # source region tag
    ("AI Country",   lambda r: r["ai_country"]),     # project country from classify.py
    ("Project Type", lambda r: r["project_type"]),
    ("Title",        lambda r: r["title"]),
    ("Summary",      lambda r: r["ai_summary"] or clean(r["summary"])),
    ("Source",       lambda r: r["url"]),
    ("Source Type",  lambda r: r["source_type"]),
    ("Language",     lambda r: r["language"]),
    ("Status",       lambda r: "New"),
]


def write_grid(ws, headers: list[str], rows: list[list]) -> None:
    """Write a header + data grid with house styling and auto column widths."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"

    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], len(str(val)) if val is not None else 0)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            MIN_WIDTH, min(MAX_WIDTH, w + 2)
        )
    # Wrap the long free-text columns (Title, Summary) for readability.
    for idx, header in enumerate(headers, start=1):
        if header in ("Title", "Summary"):
            for cell in ws[get_column_letter(idx)][1:]:
                cell.alignment = WRAP_TOP


# Sort: newest publication date first; items with no detectable date ('' /
# NULL, shown as "unknown") sink to the bottom; within a date, higher score wins.
ORDER_BY = (
    "ORDER BY (published_date IS NULL OR published_date = ''), "
    "published_date DESC, (score IS NULL), score DESC, id DESC"
)


def item_rows(conn: sqlite3.Connection, where: str = "", params: tuple = ()):
    sql = "SELECT * FROM items " + (f"WHERE {where} " if where else "") + ORDER_BY
    rows = conn.execute(sql, params).fetchall()
    return [[fn(r) for _, fn in COLUMNS] for r in rows], len(rows)


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Make sure later-added columns (ai_country, published_date) exist, so
    # export works even on a DB created before those features shipped.
    ensure_columns(conn)

    total = conn.execute("SELECT COUNT(*) FROM items").fetchone()[0]
    if total == 0:
        print("No items in intel.db yet — run collect.py first.")
        conn.close()
        return

    headers = [h for h, _ in COLUMNS]
    wb = Workbook()

    # --- ALL sheet ---
    ws_all = wb.active
    ws_all.title = "ALL"
    all_rows, n_all = item_rows(conn)
    write_grid(ws_all, headers, all_rows)

    # --- Fresh sheet: only items published within the last FRESH_DAYS days.
    # Items with no detectable date ('' -> shown "unknown") are excluded. ---
    cutoff = (date.today() - timedelta(days=FRESH_DAYS)).isoformat()
    ws_fresh = wb.create_sheet(title="Fresh")
    fresh_rows, n_fresh = item_rows(
        conn, "published_date != '' AND published_date IS NOT NULL "
        "AND published_date >= ?", (cutoff,)
    )
    write_grid(ws_fresh, headers, fresh_rows)

    # --- per-country sheets ---
    countries = [
        r[0] for r in conn.execute(
            "SELECT DISTINCT country FROM items "
            "WHERE country IS NOT NULL AND country != '' ORDER BY country"
        )
    ]
    per_country = {}
    used = {"all", "fresh", "sources", "log"}
    for c in countries:
        rows, n = item_rows(conn, "country = ?", (c,))
        title = sheet_safe(c)
        base, k = title, 2
        while title.lower() in used:
            title = f"{base[:28]}_{k}"
            k += 1
        used.add(title.lower())
        ws = wb.create_sheet(title=title)
        write_grid(ws, headers, rows)
        per_country[c] = n

    # --- SOURCES sheet (the registry) ---
    ws_src = wb.create_sheet(title="SOURCES")
    src_headers = ["Name", "Type", "Language", "Country", "Enabled", "URL"]
    src_rows = []
    if SOURCES_PATH.exists():
        data = yaml.safe_load(SOURCES_PATH.read_text(encoding="utf-8")) or {}
        for s in data.get("sources", []) or []:
            src_rows.append([
                s.get("name", ""), s.get("type", ""), s.get("language", ""),
                s.get("country", ""), "yes" if s.get("enabled") else "no",
                s.get("url", ""),
            ])
    write_grid(ws_src, src_headers, src_rows)

    # --- LOG sheet (health check: last 300 collection log rows) ---
    ws_log = wb.create_sheet(title="LOG")
    log_headers = ["Run At", "Source", "In Feed", "New", "Status", "Error"]
    log_rows = [
        list(r) for r in conn.execute(
            "SELECT run_at, source_name, items_found, items_new, status, "
            "COALESCE(error,'') FROM collection_log ORDER BY id DESC LIMIT 300"
        )
    ]
    write_grid(ws_log, log_headers, log_rows)

    conn.close()

    out = OUT_PATH
    try:
        wb.save(out)
    except PermissionError:
        # leads.xlsx is open in Excel (locked). Write a timestamped copy instead
        # so the run still succeeds; close the file to overwrite it next time.
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = OUT_PATH.with_name(f"leads_{stamp}.xlsx")
        wb.save(out)
        print(f"NOTE: {OUT_PATH.name} was locked (open in Excel) — wrote {out.name} instead.")

    generated = datetime.now(timezone.utc).isoformat(timespec="seconds")
    print(f"Wrote {out.name}  ({generated})")
    print(f"  ALL           : {n_all} rows")
    print(f"  Fresh (<= {FRESH_DAYS}d) : {n_fresh} rows  (published on/after {cutoff})")
    for c, n in per_country.items():
        print(f"  {c:<13} : {n} rows")
    print(f"  SOURCES       : {len(src_rows)} rows")
    print(f"  LOG           : {len(log_rows)} rows")


if __name__ == "__main__":
    main()
